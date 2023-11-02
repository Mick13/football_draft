#Fantasy football python program #
# python3 -m pip install openpyxl
import ctypes
import openpyxl
from openpyxl import Workbook
import os
 

"""
Drafting software for american football fictional roster.
Attatched player files contain names with aribiraty ages assigned to them. 
EXCEL export supported as well - 
to run write in termianal: # python3 -m pip install openpyxl
in the process of modifying into a full stack application
Created and completed 07/2022. Uploaded to github 11/2023
used ArrayList to support other functionalitys
"""


import ctypes  # provides low-level arrays
def make_array(n):
    return (n * ctypes.py_object)()

class ArrayList:
    def __init__(self):
        self.data_arr = make_array(1)
        self.capacity = 1
        self.n = 0

    ## functinality len(x)
    # Dunder method
    def __len__(self):
        return self.n


    def append(self, val): #amortized theta(1)
        if (self.n == self.capacity):
            self.resize(2 * self.capacity)
        self.data_arr[self.n] = val
        self.n += 1

    def resize(self, new_size):
        new_array = make_array(new_size)
        for i in range(self.n):
            new_array[i] = self.data_arr[i]
        self.data_arr = new_array
        self.capacity = new_size

    def extend(self, iter_collection):
        for elem in iter_collection:
            self.append(elem)


    def __getitem__(self, ind):
        # [x]
        if (not (-self.n <= ind <= self.n - 1)):
            raise IndexError('invalid index')
        if (ind < 0):
            ind = self.n + ind
        return self.data_arr[ind]

    def __setitem__(self, ind, val):
        if (not (-self.n <= ind <= self.n - 1)):
            raise IndexError('invalid index')
        if (ind < 0):
            ind = self.n + ind
        self.data_arr[ind] = val


    def pop(self, ind = -1):
        if (not (-self.n <= ind <= self.n - 1)):
            raise IndexError('invalid index')
        if (ind < 0):
            ind = self.n + ind
        elem = self.data_arr[ind]
        for i in range(ind+1, self.n):
            self.data_arr[i-1] = self.data_arr[i]
        self.data_arr[self.n - 1] = None
        self.n -= 1
        if (self.n < self.capacity // 4):
            self.resize(self.capacity // 2)
        return elem

    def insert(self, ind, value):
        if (not (-self.n <= ind <= self.n - 1)):
            raise IndexError('invalid index')
        if (ind < 0):
            ind = self.n + ind
        if (self.n == self.capacity):
            self.resize(2 * self.capacity)
        for j in range(self.n, ind, -1):
            self.data_arr[j] = self.data_arr[j - 1]
        self.data_arr[ind] = value
        self.n += 1

    def __repr__(self):
        data_as_strings = [str(self.data_arr[i]) for i in range(self.n)]
        return '[' + ', '.join(data_as_strings) + ']'

    def __add__(self, other):
        res = ArrayList()
        res.extend(self)
        res.extend(other)
        return res

    def __iadd__(self, other):
        self.extend(other)
        return self

    def __mul__(self, times):
        res = ArrayList()
        for i in range(times):
            res.extend(self)
        return res

    def __rmul__(self, times):
        return self * times
    

class Player:
    def __init__(self, name, age, pos):
        self.name = name
        self.age = age
        self.cost = 0
        self.pos = pos.lower()

    def __str__(self):
        return f"Name: {self.name}, Age: {self.age}, Cost: {self.cost}"

class Team:
    def __init__(self):
        self.qb = ArrayList()
        self.rb = ArrayList()
        self.wr = ArrayList()
        self.te = ArrayList()
        self.budget = 200

    def add_player(self, Player_obj, cost):
        if (self.budget - cost) <0:
            print("Warning: Insufficient budget!")
            return

        if Player_obj.pos == "qb":
            
            self.qb.append(Player_obj)
        elif Player_obj.pos == "rb":
            self.rb.append(Player_obj)

        elif Player_obj.pos == "wr":
            self.wr.append(Player_obj)

        elif Player_obj.pos == "te":
            self.te.append(Player_obj)

        Player_obj.cost = cost
        self.budget -= cost

    def to_excel(self, filename="team.xlsx"):
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "Position"
        ws["B1"] = "Player Name"
        ws["C1"] = "Cost"

        positions = ["QB", "RB", "WR", "TE"]
        position_lists = [self.qb, self.rb, self.wr, self.te]

        row = 2
        for pos, pos_list in zip(positions, position_lists):
            for player in pos_list:
                ws[f"A{row}"] = pos
                ws[f"B{row}"] = player.name
                ws[f"C{row}"] = player.cost
                row += 1

        ws[f"A{row+1}"] = "Remaining Budget"
        ws[f"B{row+1}"] = self.budget

        wb.save(filename)

#GLOBAL

player_repository = {
    'qb': ArrayList(),
    'rb': ArrayList(),
    'wr': ArrayList(),
    'te': ArrayList()
}
def load_repository():
    """
    Load players from files into the repository.
    """
    
    position_mapping = {
        "Wide Receivers": "wr",
        "Running Backs": "rb",
        "Tight Ends": "te",
        "Quarterbacks": "qb"
    }

    # Iterate over the keys of the position_mapping
    for full_position, short_position in position_mapping.items():
        filename = short_position.upper() + ".txt"
        if not os.path.exists(filename):
            print(f"File {filename} not found!")
            return

        with open(filename, 'r') as file:
            # read the first line which is the position
            actual_position = file.readline().strip()
            
            # Convert full position name to its short format
            actual_position_short = position_mapping.get(actual_position, None)

            # Check if the read position is valid
            if not actual_position_short:
                print(f"Unexpected position {actual_position} in file {filename}.")
                continue

            lines = file.readlines()
            for line in lines:

                if line.startswith("Tier") or not line:
                    continue  # skip the tier lines
                name, age = line.strip().split(',')
                player = Player(name, int(age), actual_position_short)
                player_repository[actual_position_short].append(player)


def display_rankings():
    """
    Display the rankings of all players.
    """
    for position, players in player_repository.items():
        print(f"{position.upper()} Players:")
        for i in range(len(players)):
            player = players[i]
            print(f"{i + 1}. {player.name}, Age: {player.age}")
        print()

def remove_player(position, name):
    """
    Remove a player from the repository based on the name.
    """
    players = player_repository.get(position, None)
    
    if not players:
        print(f"Position {position} not found!")
        return

    for i in range(len(players)):
        player = players.__getitem__(i)
        if player.name == name:
            players.pop(i)
            print(f"Removed {name} from {position.upper()} list.")
            return
    print(f"Player {name} not found in {position.upper()} list.")


def get_player_from_repository(name, position):
    """
    Retrieve a player object from the repository based on name and position.
    """
    players = player_repository.get(position.lower(), None)
    
    if not players:
        return None

    for i in range(len(players)):
        player = players[i]
        if player.name == name:
            return player

    return None


def main():
    
    load_repository()
    print(player_repository)
    my_team = Team()  # Create an instance of the Team class
    
    while True:
        print("\nOptions:")
        print("1: Display Rankings")
        print("2: Remove Player from Repository")
        print("3: Add Player to Team")
        print("4: Export Team to Excel")
        print("Type 'Stop' to exit.")
        
        function = input("\nEnter a function: ")
        
        if function == "1":
            display_rankings()
            
        elif function == "2":
            position = input("Enter the position of the player to be removed (e.g., qb, rb, wr, te): ")
            name = input("Enter the name of the player to be removed: ")
            remove_player(position, name)
            
        elif function == "3":
            position = input("Enter the position of the player to be added (e.g., qb, rb, wr, te): ")
            name = input("Enter the name of the player to be added: ")
            cost = float(input("Enter the cost of the player: "))
            
            player = get_player_from_repository(name, position)
            if player:
                my_team.add_player(player, cost)
            else:
                print(f"No player named {name} found in {position}.")
            
        # elif function == "4":
        #     pass
            
        elif function == "4":
            filename = input("Enter the filename (default: team.xlsx): ")
            if not filename:
                filename = "team.xlsx"
            my_team.to_excel(filename)
            
        elif function.lower() == "stop":
            break
            
        else:
            print("Invalid option, please try again.")

if __name__ == "__main__":
    main()
